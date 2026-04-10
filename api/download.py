"""
Serverless функция для генерации многолистового Excel отчета
URL: /api/download
С ПОДДЕРЖКОЙ ПАГИНАЦИИ - получает ВСЕ записи
"""

from http.server import BaseHTTPRequestHandler
import json
import re
import requests
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime, date
import io

# Конфигурация
KOBO_API_TOKEN = '929c90ea6bbce9e24789c10b2eb9740e3352d859'
ASSET_ID = 'aCE5fencfcUpVhvCRdCoxc'

# ---- Фильтр Шымкента: учитываем визиты только с 26 марта 2026 ----
SHYMKENT_MIN_DATE = date(2026, 3, 26)
DATE_RE = re.compile(r'(\d{4}-\d{2}-\d{2})')


def parse_visit_date(raw):
    """Парсит дату из строки group_xn8xb93/date в datetime.date или None."""
    if not raw:
        return None
    m = DATE_RE.search(str(raw))
    if not m:
        return None
    try:
        return datetime.strptime(m.group(1), '%Y-%m-%d').date()
    except ValueError:
        return None


def get_validation_status(record):
    """Возвращает 'Валидирована' / 'Не валидирована' / 'Без статуса'."""
    vs = record.get('_validation_status')
    if isinstance(vs, dict):
        uid = vs.get('uid', '')
        if uid == 'validation_status_approved':
            return 'Валидирована'
        if uid == 'validation_status_not_approved':
            return 'Не валидирована'
    return 'Без статуса'

# Маппинг кодов
CITY_CODES = {
    '710000000': 'г. Астана',
    '750000000': 'г. Алматы',
    '790000000': 'г. Шымкент',
    '151000000': 'Актобе Г.А.'
}

RESULT_CODES = {
    '1': 'Контакт установлен',
    '2': 'Неконтакт',
    '3': 'Недоступный адрес',
    '4': 'Языковой барьер',
    '5': 'Отказ',
    '6': 'Другое'
}

QUOTAS = {
    'г. Астана': {'total': 800, 'employed': 608, 'self_employed': 192, 'peo_count': 32},
    'г. Алматы': {'total': 975, 'employed': 741, 'self_employed': 234, 'peo_count': 39},
    'г. Шымкент': {'total': 700, 'employed': 532, 'self_employed': 168, 'peo_count': 28},
    'Актобе Г.А.': {'total': 525, 'employed': 399, 'self_employed': 126, 'peo_count': 21}
}

def fetch_kobo_data():
    """
    Загрузка ВСЕХ данных из Kobo с пагинацией
    """
    headers = {'Authorization': f'Token {KOBO_API_TOKEN}'}
    
    all_results = []
    url = f"https://kf.kobotoolbox.org/api/v2/assets/{ASSET_ID}/data.json?limit=100"
    
    while url:
        response = requests.get(url, headers=headers, timeout=30)
        response.raise_for_status()
        
        data = response.json()
        results = data.get('results', [])
        all_results.extend(results)
        
        # Следующая страница
        url = data.get('next')
        
        # Безопасность: останавливаемся после 5000 записей
        if len(all_results) > 5000:
            break
    
    return all_results

def process_data(records):
    """Обработка данных"""
    processed = []
    
    for record in records:
        # Правильные названия полей
        city_code = record.get('city', '')
        city = CITY_CODES.get(str(city_code), f'Неизвестно ({city_code})')
        
        # Дата и время
        date_raw = record.get('group_xn8xb93/date', '')
        time_raw = record.get('group_xn8xb93/time', '')
        time_clean = str(time_raw).split('+')[0].split('.')[0] if time_raw else ''
        
        # Парсим дату визита для фильтра
        visit_date = parse_visit_date(date_raw)
        
        # Статус валидации
        validation = get_validation_status(record)
        
        # Результат
        result_code = record.get('group_ip3jm92/result', '')
        result = RESULT_CODES.get(str(result_code), f'Неизвестно ({result_code})')
        
        # Готовность и согласие
        willingness = record.get('willingness', '')
        consent = record.get('consent', '')
        q08 = record.get('q08_survey2', '')
        
        # Определяем завершенность
        is_completed = (
            str(willingness) == '1' and
            str(consent) == '1'
        )
        
        # Категория
        if is_completed and str(q08).strip():
            q08_val = str(q08).strip()
            if q08_val == '1':
                category = 'Наемный работник'
            elif q08_val in ['2', '3', '4', '5']:
                category = 'Самозанятый/ИП'
            else:
                category = 'Другое'
        else:
            category = 'Другое'
        
        # Язык респондента
        language = record.get('group_xl1fx65/lang_resp', '')
        
        # Проверка на контакт
        is_contact = str(result_code) == '1'
        
        processed.append({
            'date': date_raw,
            'time': time_clean,
            'city': city,
            'peo': record.get('group_xn8xb93/PEO', ''),
            'segment': record.get('group_xn8xb93/segment_num', ''),
            'interviewer': record.get('group_xn8xb93/int_name', ''),
            'result': result,
            'category': category,
            'language': language,
            'attempt': record.get('group_xn8xb93/attempt', ''),
            'is_completed': is_completed,
            'is_contact': is_contact,
            'validation': validation,
            'visit_date': visit_date,
        })
    
    # --- Фильтр Шымкента: исключаем визиты до 26 марта 2026 ---
    def keep(r):
        if r['city'] != 'г. Шымкент':
            return True
        if r['visit_date'] is None:
            return False
        return r['visit_date'] >= SHYMKENT_MIN_DATE
    
    processed = [r for r in processed if keep(r)]
    
    return processed

def create_dashboard_sheet(wb, processed_data):
    """Создание листа Dashboard"""
    ws = wb.create_sheet("Dashboard", 0)
    
    # Стили
    header_font = Font(bold=True, size=14, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    # Заголовок
    ws['A1'] = 'Мониторинг полевого опроса ЕНПФ'
    ws['A1'].font = Font(bold=True, size=16)
    
    # Основные метрики
    total_visits = len(processed_data)
    completed = sum(1 for r in processed_data if r['is_completed'])
    contacts = sum(1 for r in processed_data if r['is_contact'])
    
    total_quota = sum(q['total'] for q in QUOTAS.values())
    rr = round((completed / total_visits * 100) if total_visits > 0 else 0, 1)
    cr = round((contacts / total_visits * 100) if total_visits > 0 else 0, 1)
    
    ws['A4'] = 'Метрика'
    ws['B4'] = 'Значение'
    ws['A4'].font = header_font
    ws['A4'].fill = header_fill
    ws['B4'].font = header_font
    ws['B4'].fill = header_fill
    
    metrics = [
        ('Всего визитов', total_visits),
        ('Завершено опросов', completed),
        ('Общая квота', total_quota),
        ('Прогресс по квоте (%)', f"{round((completed/total_quota*100) if total_quota > 0 else 0, 1)}%"),
        ('Response Rate (%)', f"{rr}%"),
        ('Contact Rate (%)', f"{cr}%")
    ]
    
    for idx, (metric, value) in enumerate(metrics, start=5):
        ws[f'A{idx}'] = metric
        ws[f'B{idx}'] = value
    
    # По городам
    ws['A13'] = 'Прогресс по городам'
    ws['A13'].font = Font(bold=True, size=12)
    
    ws['A14'] = 'Город'
    ws['B14'] = 'Завершено'
    ws['C14'] = 'Квота'
    ws['D14'] = 'Прогресс (%)'
    ws['E14'] = 'Наемных'
    ws['F14'] = 'Самозанятых'
    ws['G14'] = 'Валидированы'
    ws['H14'] = 'Не валидированы'
    ws['I14'] = 'Без статуса'
    
    for col in ['A14', 'B14', 'C14', 'D14', 'E14', 'F14', 'G14', 'H14', 'I14']:
        ws[col].font = header_font
        ws[col].fill = header_fill
    
    row = 15
    for city, quota in QUOTAS.items():
        city_records = [r for r in processed_data if r['city'] == city]
        city_completed = sum(1 for r in city_records if r['is_completed'])
        city_employed = sum(1 for r in city_records if r['category'] == 'Наемный работник')
        city_self = sum(1 for r in city_records if r['category'] == 'Самозанятый/ИП')
        city_approved = sum(1 for r in city_records if r['validation'] == 'Валидирована')
        city_not_approved = sum(1 for r in city_records if r['validation'] == 'Не валидирована')
        city_no_status = sum(1 for r in city_records if r['validation'] == 'Без статуса')
        progress = round((city_completed / quota['total'] * 100) if quota['total'] > 0 else 0, 1)
        
        ws[f'A{row}'] = city
        ws[f'B{row}'] = city_completed
        ws[f'C{row}'] = quota['total']
        ws[f'D{row}'] = f"{progress}%"
        ws[f'E{row}'] = f"{city_employed}/{quota['employed']}"
        ws[f'F{row}'] = f"{city_self}/{quota['self_employed']}"
        ws[f'G{row}'] = city_approved
        ws[f'H{row}'] = city_not_approved
        ws[f'I{row}'] = city_no_status
        row += 1
    
    # Автоширина
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

def create_peo_sheet(wb, processed_data):
    """Создание листа по ПЕО"""
    ws = wb.create_sheet("Polling Station")
    
    headers = ['Город', 'ПЕО', 'Интервьюер', 'Всего визитов', 'Завершено', 'RR (%)', 'CR (%)']
    ws.append(headers)
    
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    
    # Группировка
    peo_stats = {}
    for record in processed_data:
        key = (record['city'], record['peo'], record['interviewer'])
        if key not in peo_stats:
            peo_stats[key] = {'visits': 0, 'completed': 0, 'contacts': 0}
        
        peo_stats[key]['visits'] += 1
        if record['is_completed']:
            peo_stats[key]['completed'] += 1
        if record['is_contact']:
            peo_stats[key]['contacts'] += 1
    
    # Данные
    for (city, peo, interviewer), stats in sorted(peo_stats.items()):
        rr = round((stats['completed'] / stats['visits'] * 100) if stats['visits'] > 0 else 0, 1)
        cr = round((stats['contacts'] / stats['visits'] * 100) if stats['visits'] > 0 else 0, 1)
        
        ws.append([city, peo, interviewer, stats['visits'], stats['completed'], f"{rr}%", f"{cr}%"])
    
    # Автоширина
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

def create_interviewer_sheet(wb, processed_data):
    """Создание листа по интервьюерам"""
    ws = wb.create_sheet("Enumerator & Supervisor")
    
    headers = ['Интервьюер', 'Город', 'Всего визитов', 'Завершено', 'Неконтакты', 'Отказы', 'RR (%)', 'CR (%)']
    ws.append(headers)
    
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    
    # Группировка
    int_stats = {}
    for record in processed_data:
        key = (record['interviewer'], record['city'])
        if key not in int_stats:
            int_stats[key] = {'visits': 0, 'completed': 0, 'contacts': 0, 'refusals': 0}
        
        int_stats[key]['visits'] += 1
        if record['is_completed']:
            int_stats[key]['completed'] += 1
        if record['is_contact']:
            int_stats[key]['contacts'] += 1
        if 'Отказ' in record['result']:
            int_stats[key]['refusals'] += 1
    
    # Данные
    for (interviewer, city), stats in sorted(int_stats.items()):
        non_contacts = stats['visits'] - stats['contacts']
        rr = round((stats['completed'] / stats['visits'] * 100) if stats['visits'] > 0 else 0, 1)
        cr = round((stats['contacts'] / stats['visits'] * 100) if stats['visits'] > 0 else 0, 1)
        
        ws.append([interviewer, city, stats['visits'], stats['completed'], non_contacts, stats['refusals'], f"{rr}%", f"{cr}%"])
    
    # Автоширина
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

def create_raw_data_sheet(wb, processed_data):
    """Создание листа Raw Data"""
    ws = wb.create_sheet("Raw Data")
    
    headers = ['Дата', 'Время', 'Город', 'ПЕО', 'Сегмент', 'Интервьюер', 'Результат', 'Категория', 'Язык', 'Попытка', 'Статус валидации']
    ws.append(headers)
    
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    
    # Данные
    for record in processed_data:
        ws.append([
            record['date'],
            record['time'],
            record['city'],
            record['peo'],
            record['segment'],
            record['interviewer'],
            record['result'],
            record['category'],
            record['language'],
            record['attempt'],
            record['validation']
        ])
    
    # Автоширина
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

def create_excel_report(processed_data):
    """Создание Excel отчета"""
    wb = Workbook()
    
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    create_dashboard_sheet(wb, processed_data)
    create_peo_sheet(wb, processed_data)
    create_interviewer_sheet(wb, processed_data)
    create_raw_data_sheet(wb, processed_data)
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output.getvalue()

class handler(BaseHTTPRequestHandler):
    def do_GET(self):
        try:
            records = fetch_kobo_data()
            processed_data = process_data(records)
            excel_bytes = create_excel_report(processed_data)
            
            filename = f"ENPF_Survey_Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            
            self.send_response(200)
            self.send_header('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            self.send_header('Content-Disposition', f'attachment; filename="{filename}"')
            self.send_header('Access-Control-Allow-Origin', '*')
            self.end_headers()
            
            self.wfile.write(excel_bytes)
            
        except Exception as e:
            self.send_response(500)
            self.send_header('Content-type', 'application/json')
            self.end_headers()
            
            error_response = {
                'error': str(e),
                'message': 'Ошибка генерации отчета'
            }
            self.wfile.write(json.dumps(error_response, ensure_ascii=False).encode('utf-8'))
